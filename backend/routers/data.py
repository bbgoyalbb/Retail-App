"""
Data router.
"""
from fastapi import APIRouter, Depends, HTTPException, Query, Request, UploadFile, File, Body
from fastapi.responses import StreamingResponse, FileResponse
from typing import Optional, List, Dict, Any
from datetime import datetime, timezone, date
import uuid
import re
import os
import logging
logger = logging.getLogger(__name__)
from bson import ObjectId
from .deps import db, get_current_user_dep
from data_quality import round_money, determine_payment_status, build_payment_mode_label
import auth as auth_module
from auth import audit_log
from .models import ADDON_ITEMS, ARTICLE_TYPES, ItemUpdateRequest, PAYMENT_MODES, TAILORING_RATES, DEFAULT_SETTINGS, merge_settings
import io, json, shutil
from pathlib import Path
from data_quality import generate_data_audit as dq_generate_data_audit, normalize_low_risk_data as dq_normalize_low_risk_data, repair_high_risk_data as dq_repair_high_risk_data
ROOT_DIR = Path(__file__).parent.parent

def safe_float(v):
    if v is None or str(v).strip() in ("N/A", "", "None"):
        return 0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0

def safe_str(v):
    if v is None:
        return "N/A"
    return str(v).strip()

def safe_date(v):
    if v and hasattr(v, 'strftime'):
        return v.strftime("%Y-%m-%d")
    return "N/A"

router = APIRouter()

@router.post("/import/excel")
async def import_excel(
    file: UploadFile = File(...),
    mode: str = "replace",
    current_user: dict = Depends(get_current_user_dep)
):
    # Restrict to admin only
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    if not file.filename.endswith(('.xlsm', '.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsm or .xlsx)")

    try:
        import openpyxl
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)

        items_count = 0
        advances_count = 0
        items = []
        advances = []

        if 'Item Details' in wb.sheetnames:
            ws = wb['Item Details']
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row or None, values_only=True):
                if not any(cell is not None for cell in row):
                    break
                if not row[0]:
                    continue

                item = {
                    "id": str(uuid.uuid4()),
                    "date": safe_date(row[0]),
                    "name": safe_str(row[1]),
                    "ref": safe_str(row[2]),
                    "barcode": safe_str(row[3]),
                    "price": safe_float(row[4]),
                    "qty": safe_float(row[5]),
                    "discount": safe_float(row[6]),
                    "fabric_amount": safe_float(row[7]),
                    "tailoring_status": safe_str(row[8]),
                    "article_type": safe_str(row[9]),
                    "order_no": safe_str(row[10]),
                    "delivery_date": safe_date(row[11]),
                    "tailoring_amount": safe_float(row[12]),
                    "embroidery_status": safe_str(row[13]),
                    "embroidery_amount": safe_float(row[14]),
                    "addon_desc": safe_str(row[15]),
                    "addon_amount": safe_float(row[16]),
                    "fabric_pay_mode": safe_str(row[17]),
                    "fabric_pay_date": safe_date(row[18]),
                    "fabric_pending": safe_float(row[19]),
                    "fabric_received": safe_float(row[20]),
                    "labour_amount": safe_float(row[21]),
                    "labour_paid": safe_str(row[22]),
                    "labour_pay_date": safe_date(row[23]),
                    "tailoring_pay_mode": safe_str(row[24]),
                    "tailoring_pay_date": safe_date(row[25]),
                    "tailoring_received": safe_float(row[26]),
                    "tailoring_pending": safe_float(row[27]),
                    "embroidery_pay_mode": safe_str(row[28]),
                    "embroidery_pay_date": safe_date(row[29]),
                    "embroidery_received": safe_float(row[30]),
                    "embroidery_pending": safe_float(row[31]),
                    "addon_pay_mode": safe_str(row[32]),
                    "addon_pay_date": safe_date(row[33]),
                    "addon_received": safe_float(row[34]),
                    "addon_pending": safe_float(row[35]),
                    "karigar": safe_str(row[36]) if len(row) > 36 else "N/A",
                    "emb_labour_amount": safe_float(row[37]) if len(row) > 37 else 0,
                    "emb_labour_paid": safe_str(row[38]) if len(row) > 38 else "N/A",
                    "emb_labour_date": safe_date(row[39]) if len(row) > 39 else "N/A",
                    "emb_labour_payment_mode": safe_str(row[40]) if len(row) > 40 else "N/A",
                    # Read tally columns if present in Excel, otherwise default to False
                    "tally_fabric": bool(row[41]) if len(row) > 41 and row[41] not in (None, "", "N/A") else False,
                    "tally_tailoring": bool(row[42]) if len(row) > 42 and row[42] not in (None, "", "N/A") else False,
                    "tally_embroidery": bool(row[43]) if len(row) > 43 and row[43] not in (None, "", "N/A") else False,
                    "tally_addon": bool(row[44]) if len(row) > 44 and row[44] not in (None, "", "N/A") else False,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                }
                items.append(item)

            if items:
                items_count = len(items)

        if 'Advances' in wb.sheetnames:
            ws2 = wb['Advances']
            for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row or None, values_only=True):
                if not any(cell is not None for cell in row):
                    break
                if not row[0]:
                    continue
                dt = row[0]
                date_str = dt.strftime("%Y-%m-%d") if hasattr(dt, 'strftime') else str(dt)[:10]
                adv = {
                    "id": str(uuid.uuid4()),
                    "date": date_str,
                    "name": str(row[1]).strip() if row[1] else "",
                    "ref": str(row[2]).strip() if row[2] else "",
                    "amount": float(row[3]) if row[3] else 0,
                    "mode": str(row[4]).strip() if row[4] else "",
                    # Read tally column if present in Excel, otherwise default to False
                    "tally": bool(row[5]) if len(row) > 5 and row[5] not in (None, "", "N/A") else False,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                }
                advances.append(adv)

            if advances:
                advances_count = len(advances)

        if mode == "replace":
            # Atomic replace using transaction (same pattern as restore)
            try:
                async with await db.client.start_session() as session:
                    async with session.start_transaction():
                        await db.items.delete_many({}, session=session)
                        await db.advances.delete_many({}, session=session)
                        if items:
                            await db.items.insert_many(items, session=session)
                        if advances:
                            await db.advances.insert_many(advances, session=session)
            except Exception as txn_err:
                # Fallback for standalone MongoDB (no replica set)
                logger.warning(f"Transaction not available for import: {txn_err}. Falling back to non-atomic replace.")
                await db.items.delete_many({})
                await db.advances.delete_many({})
                if items:
                    await db.items.insert_many(items)
                if advances:
                    await db.advances.insert_many(advances)
        else:
            # Append mode - no deletion needed
            if items:
                await db.items.insert_many(items)
            if advances:
                await db.advances.insert_many(advances)

        return {
            "message": f"Import successful! {items_count} items and {advances_count} advances imported.",
            "items_count": items_count,
            "advances_count": advances_count,
        }
    except Exception as e:
        logger.error(f"Import error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Import failed. Check the file format and try again.")

# ==========================================
# EXCEL EXPORT
# ==========================================

@router.get("/export/excel")
async def export_excel(current_user: dict = Depends(get_current_user_dep)):
    # Restrict to admin only
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    # Items sheet
    ws = wb.active
    ws.title = "Item Details"
    headers = [
        "Date", "Name", "Ref.", "Items", "Price", "Qty", "Discount", "Fabric Amount",
        "Tailoring?", "Article Type", "Order No.", "Delivery Date", "Tailoring Amount",
        "Embroidery?", "Embroidery Amount", "Add-on", "Add-on Amount",
        "Fabric Payment Mode", "Fabric Payment Date", "Fabric Pending Balance", "Fabric Payment Received",
        "Labour Amount", "Labour Paid?", "Labour Payment Date",
        "Tailoring Payment Mode", "Tailoring Payment Date", "Tailoring Payment Received", "Tailoring Pending Balance",
        "Embroidery Payment Mode", "Embroidery Payment Date", "Embroidery Payment Received", "Embroidery Pending Balance",
        "Add-On Payment Mode", "Add-On Payment Date", "Add-On Payment Received", "Add-On Pending Balance",
        "Karigar?", "Emb Labour Amount", "Emb Labour Paid?", "Emb Labour Date", "Emb Labour Payment Mode",
    ]

    header_fill = PatternFill(start_color="C86B4D", end_color="C86B4D", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    items = await db.items.find({}, {"_id": 0}).sort("date", 1).to_list(10000)
    fields = [
        "date", "name", "ref", "barcode", "price", "qty", "discount", "fabric_amount",
        "tailoring_status", "article_type", "order_no", "delivery_date", "tailoring_amount",
        "embroidery_status", "embroidery_amount", "addon_desc", "addon_amount",
        "fabric_pay_mode", "fabric_pay_date", "fabric_pending", "fabric_received",
        "labour_amount", "labour_paid", "labour_pay_date",
        "tailoring_pay_mode", "tailoring_pay_date", "tailoring_received", "tailoring_pending",
        "embroidery_pay_mode", "embroidery_pay_date", "embroidery_received", "embroidery_pending",
        "addon_pay_mode", "addon_pay_date", "addon_received", "addon_pending",
        "karigar", "emb_labour_amount", "emb_labour_paid", "emb_labour_date", "emb_labour_payment_mode",
    ]

    for row_idx, item in enumerate(items, 2):
        for col_idx, field in enumerate(fields, 1):
            ws.cell(row=row_idx, column=col_idx, value=item.get(field, ""))

    # Advances sheet
    ws2 = wb.create_sheet("Advances")
    adv_headers = ["Advance Payment Date", "Name", "Ref", "Advance Payment Amount", "Advance Payment Mode"]
    for col, header in enumerate(adv_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    advances = await db.advances.find({}, {"_id": 0}).sort("date", 1).to_list(5000)
    for row_idx, adv in enumerate(advances, 2):
        ws2.cell(row=row_idx, column=1, value=adv.get("date", ""))
        ws2.cell(row=row_idx, column=2, value=adv.get("name", ""))
        ws2.cell(row=row_idx, column=3, value=adv.get("ref", ""))
        ws2.cell(row=row_idx, column=4, value=adv.get("amount", 0))
        ws2.cell(row=row_idx, column=5, value=adv.get("mode", ""))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"retail_book_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ==========================================
# DATABASE BACKUP & RESTORE
# ==========================================

@router.get("/backup")
async def backup_database(current_user: dict = Depends(get_current_user_dep)):
    # Restrict to admin only
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    items = await db.items.find({}, {"_id": 0}).to_list(50000)
    advances = await db.advances.find({}, {"_id": 0}).to_list(10000)

    backup_data = {
        "version": "1.0",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "items_count": len(items),
        "advances_count": len(advances),
        "items": items,
        "advances": advances,
    }

    buffer = io.BytesIO(json.dumps(backup_data, indent=2, default=str).encode('utf-8'))
    buffer.seek(0)

    filename = f"retail_backup_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
    return StreamingResponse(
        buffer,
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.post("/restore")
async def restore_database(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user_dep)
):
    # Restrict to admin only
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    if not file.filename.endswith('.json'):
        raise HTTPException(status_code=400, detail="Please upload a .json backup file")

    try:
        contents = await file.read()
        backup_data = json.loads(contents.decode('utf-8'))

        # ===== VALIDATION PHASE =====
        # Validate backup structure before touching database
        if "items" not in backup_data or "advances" not in backup_data:
            raise HTTPException(status_code=400, detail="Invalid backup file format: missing 'items' or 'advances'")

        items = backup_data["items"]
        advances = backup_data["advances"]

        if not isinstance(items, list) or not isinstance(advances, list):
            raise HTTPException(status_code=400, detail="Invalid backup file format: 'items' and 'advances' must be lists")
        
        if len(items) == 0 and len(advances) == 0:
            raise HTTPException(status_code=400, detail="Backup file contains no data")

        # Validate sample of items have required fields
        required_item_fields = {"id", "ref", "name", "date", "barcode"}
        for i, item in enumerate(items[:10]):
            if not isinstance(item, dict):
                raise HTTPException(status_code=400, detail=f"Invalid item at index {i}: not an object")
            missing = required_item_fields - set(item.keys())
            if missing:
                raise HTTPException(status_code=400, detail=f"Item at index {i} missing required fields: {missing}")

        # Validate advances have required fields
        required_advance_fields = {"id", "ref", "name", "amount"}
        for i, adv in enumerate(advances[:10]):
            if not isinstance(adv, dict):
                raise HTTPException(status_code=400, detail=f"Invalid advance at index {i}: not an object")
            missing = required_advance_fields - set(adv.keys())
            if missing:
                raise HTTPException(status_code=400, detail=f"Advance at index {i} missing required fields: {missing}")

        # Get counts for response
        items_count = len(items)
        advances_count = len(advances)

        # Audit log only after validation passes — confirmed it is a real restore attempt
        await audit_log(db, "restore", current_user, "database", "full", {"filename": file.filename, "items": items_count, "advances": advances_count})

        # ===== RESTORE PHASE =====
        # Only delete after full validation passes
        # Wrap in MongoDB transaction for atomicity (requires replica set)
        try:
            # Try to use a session for transaction (requires MongoDB replica set)
            async with await db.client.start_session() as session:
                async with session.start_transaction():
                    await db.items.delete_many({}, session=session)
                    await db.advances.delete_many({}, session=session)
                    if items:
                        await db.items.insert_many(items, session=session)
                    if advances:
                        await db.advances.insert_many(advances, session=session)
        except Exception as txn_err:
            # Fallback for standalone MongoDB (no replica set) - log warning and proceed without transaction
            logger.warning(f"MongoDB transaction failed (expected for standalone MongoDB): {txn_err}. Proceeding with non-atomic restore.")
            await db.items.delete_many({})
            await db.advances.delete_many({})
            if items:
                await db.items.insert_many(items)
            if advances:
                await db.advances.insert_many(advances)

        return {
            "message": f"Restore successful! {items_count} items and {advances_count} advances restored.",
            "items_count": items_count,
            "advances_count": advances_count,
        }
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid JSON file")
    except HTTPException:
        raise  # Re-raise HTTP exceptions as-is
    except Exception as e:
        logger.error(f"Restore error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Restore failed. Check server logs for details.")

@router.get("/db/stats")
async def get_db_stats(current_user: dict = Depends(get_current_user_dep)):
    items_count = await db.items.count_documents({})
    advances_count = await db.advances.count_documents({})
    return {
        "items_count": items_count,
        "advances_count": advances_count,
        "db_name": os.environ.get('DB_NAME', 'unknown'),
    }

@router.get("/db/audit")
async def get_db_audit(limit: int = 100, current_user: dict = Depends(get_current_user_dep)):
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only admins can access the data audit")
    safe_limit = max(1, min(limit, 500))
    return await dq_generate_data_audit(db, safe_limit)

@router.post("/db/normalize")
async def normalize_db_data(limit: int = 100, current_user: dict = Depends(get_current_user_dep)):
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only admins can normalize data")
    safe_limit = max(1, min(limit, 500))
    return await dq_normalize_low_risk_data(db, safe_limit)

@router.post("/db/repair")
async def repair_db_data(limit: int = 100, current_user: dict = Depends(get_current_user_dep)):
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only admins can repair data")
    safe_limit = max(1, min(limit, 500))
    return await dq_repair_high_risk_data(db, safe_limit)

# ==========================================
# SETTINGS (authenticated)
# ==========================================


